#!/usr/bin/env python3
"""
Build operator → EDGAR parent mapping v4.
Uses GEOT entity_parent_lookup + optimized matching against Supabase operators.
"""

import json
import re
import urllib.request
from urllib.request import Request
from collections import defaultdict

SUPABASE_URL = "https://temtptsfiksixxhbigkg.supabase.co"
SUPABASE_KEY = "os.environ.get('SUPABASE_KEY')"
H = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

DATA_DIR = "/home/openclaw/.openclaw/workspace/varro/data"
GEOT_PATH = f"{DATA_DIR}/raw/gem-ownership/geot_entity_parent_lookup.json"
OUTPUT_PATH = f"{DATA_DIR}/operator-parent-mapping-v4.json"

# EDGAR parent normalization map
EDGAR_PATTERNS = {
    "ExxonMobil": [
        "exxon mobil corp", "exxonmobil", "exxon mobil corporation",
    ],
    "Shell": [
        "shell plc", "shell p.l.c",
    ],
    "Chevron": [
        "chevron corp", "chevron corporation", "texaco inc",
    ],
    "BP": [
        "bp plc", "bp p.l.c",
    ],
    "Equinor": [
        "equinor asa", "statoil asa",
    ],
    "TotalEnergies": [
        "totalenergies se", "total se", "total s.a", "totalenergies",
    ],
    "ConocoPhillips": [
        "conocophillips corp", "conocophillips corporation",
    ],
    "Occidental": [
        "occidental petroleum corp", "occidental petroleum corporation",
        "oxy usa",
    ],
    "Devon": [
        "devon energy corp", "devon energy corporation", "devon energy",
    ],
    "Pioneer": [
        "pioneer natural resources", "pioneer natural",
    ],
}

# Direct keyword fallbacks for operators not in GEOT
DIRECT_KEYWORDS = {
    "ExxonMobil": ["exxon", "mobil oil", "mobil corp", "esso"],
    "Shell": ["shell oil", "shell offshore", "shell western", "shell deep", "shell gulf"],
    "Chevron": ["chevron", "texaco", "unocal", "caltex"],
    "BP": ["bp exploration", "bp america", "bp energy", "arco alaska", "amoco", "atlantic richfield"],
    "Equinor": ["equinor", "statoil"],
    "TotalEnergies": ["total e&p", "total petroleum", "fina oil", "totalenergies"],
    "ConocoPhillips": ["conocophillips", "conoco inc", "phillips petroleum", "burlington resources"],
    "Occidental": ["occidental", "oxy usa", "anadarko"],
    "Devon": ["devon energy", "santa fe"],
    "Pioneer": ["pioneer natural", "mesa petroleum"],
}

def gem_parents_to_edgar(gem_parents_str: str) -> str | None:
    if not gem_parents_str:
        return None
    gp_lower = gem_parents_str.lower()
    for edgar, patterns in EDGAR_PATTERNS.items():
        for pat in patterns:
            if pat in gp_lower:
                return edgar
    return None

def fetch_all_operators():
    operators = []
    page_size = 1000
    offset = 0
    while True:
        url = f"{SUPABASE_URL}/rest/v1/operators?select=id,name&limit={page_size}&offset={offset}"
        req = Request(url, headers=H)
        batch = json.loads(urllib.request.urlopen(req).read())
        operators.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return operators

STOPWORDS = {'corp', 'corporation', 'company', 'inc', 'ltd', 'llc', 'lp',
             'the', 'and', 'oil', 'gas', 'energy', 'petroleum', 'resources',
             'operations', 'operating', 'exploration', 'production', 'services',
             'international', 'north', 'south', 'east', 'west', 'american',
             'offshore', 'onshore', 'holdings', 'holding', 'group', 'plc'}

def tokenize(name: str) -> set:
    words = re.sub(r'[^a-z0-9\s]', ' ', name.lower()).split()
    return {w for w in words if len(w) > 4 and w not in STOPWORDS}

def build_token_index(geot_lookup: dict) -> dict:
    """Build inverted index: token -> list of geot_keys"""
    index = defaultdict(list)
    for key in geot_lookup:
        for token in tokenize(key):
            index[token].append(key)
    return index

def main():
    print("=== Operator → EDGAR Parent Mapping v4 ===\n")
    
    print("Loading GEOT entity parent lookup...")
    with open(GEOT_PATH) as f:
        geot_lookup = json.load(f)
    print(f"  {len(geot_lookup):,} GEOT entities loaded")
    
    # Supplement with xlsx extra name fields
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            f'{DATA_DIR}/raw/gem-ownership/geot.xlsx',
            read_only=True, data_only=True)
        ws = wb['All Entities']
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        print(f"  XLSX headers: {headers}")
        extra_xlsx = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            gem_parents = record.get('Gem parents') or record.get('gem_parents')
            if not gem_parents:
                continue
            for field in ['Name', 'Full Name', 'Name Local', 'Name Other', 'Abbreviation']:
                val = record.get(field)
                if val and isinstance(val, str):
                    key = val.lower().strip()
                    if key not in geot_lookup and len(key) >= 4:
                        geot_lookup[key] = gem_parents
                        extra_xlsx += 1
        print(f"  Added {extra_xlsx} entries from XLSX")
        wb.close()
    except Exception as e:
        print(f"  XLSX note: {e}")
    
    print(f"Total GEOT lookup size: {len(geot_lookup):,}")
    
    # Build token index for fast token-overlap matching
    print("\nBuilding token index...")
    token_index = build_token_index(geot_lookup)
    print(f"  Token index: {len(token_index):,} unique tokens")
    
    # Fetch operators
    print("\nFetching operators from Supabase...")
    operators = fetch_all_operators()
    print(f"  Total operators: {len(operators):,}")
    
    mapping = {}
    stats = defaultdict(int)
    edgar_breakdown = defaultdict(int)
    
    for op in operators:
        op_id = op["id"]
        op_name = (op["name"] or "").strip()
        if not op_name:
            stats["no_name"] += 1
            continue
        
        op_lower = op_name.lower()
        edgar_parent = None
        gem_parents_val = None
        method = None
        confidence = "low"
        
        # 1. Exact match in GEOT
        if op_lower in geot_lookup:
            gem_parents_val = geot_lookup[op_lower]
            edgar_parent = gem_parents_to_edgar(gem_parents_val)
            if edgar_parent:
                method = "geot_exact"
                confidence = "high"
        
        # 2. Substring: GEOT key in operator name (prefer longer keys)
        if not edgar_parent:
            best_len = 5  # min 6 chars
            for geot_key, geot_val in geot_lookup.items():
                if len(geot_key) > best_len and geot_key in op_lower:
                    candidate = gem_parents_to_edgar(geot_val)
                    if candidate:
                        best_len = len(geot_key)
                        gem_parents_val = geot_val
                        edgar_parent = candidate
                        method = "geot_substring"
                        confidence = "high"
        
        # 3. Substring: operator name in GEOT key
        if not edgar_parent and len(op_lower) >= 6:
            for geot_key, geot_val in geot_lookup.items():
                if op_lower in geot_key and len(op_lower) >= 6:
                    candidate = gem_parents_to_edgar(geot_val)
                    if candidate:
                        gem_parents_val = geot_val
                        edgar_parent = candidate
                        method = "geot_substring_rev"
                        confidence = "medium"
                        break
        
        # 4. Token overlap (fast via index)
        if not edgar_parent:
            op_tokens = tokenize(op_name)
            if len(op_tokens) >= 1:
                candidate_counts = defaultdict(int)
                for tok in op_tokens:
                    for geot_key in token_index.get(tok, []):
                        candidate_counts[geot_key] += 1
                # Find best with 2+ overlap that has an EDGAR parent
                best_overlap = 1  # need > 1
                for geot_key, cnt in candidate_counts.items():
                    if cnt > best_overlap:
                        candidate = gem_parents_to_edgar(geot_lookup.get(geot_key))
                        if candidate:
                            best_overlap = cnt
                            gem_parents_val = geot_lookup[geot_key]
                            edgar_parent = candidate
                            method = "geot_token"
                            confidence = "low"
        
        # 5. Direct keyword fallback
        if not edgar_parent:
            for edgar, keywords in DIRECT_KEYWORDS.items():
                for kw in keywords:
                    if kw in op_lower:
                        edgar_parent = edgar
                        method = "keyword_fallback"
                        confidence = "medium"
                        break
                if edgar_parent:
                    break
        
        if edgar_parent:
            mapping[op_id] = {
                "operator_name": op_name,
                "edgar_parent": edgar_parent,
                "gem_parents": gem_parents_val,
                "match_method": method,
                "confidence": confidence
            }
            stats[method] += 1
            edgar_breakdown[edgar_parent] += 1
        else:
            stats["no_match"] += 1
    
    total_matched = len(mapping)
    total_operators = len(operators)
    coverage_pct = round(total_matched / total_operators * 100, 1) if total_operators else 0
    
    output = {
        "version": "4.0",
        "total_operators": total_operators,
        "matched": total_matched,
        "coverage_pct": coverage_pct,
        "stats": dict(stats),
        "edgar_breakdown": dict(edgar_breakdown),
        "mapping": mapping
    }
    
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"\n=== Results ===")
    print(f"Operators: {total_operators:,}")
    print(f"Matched to EDGAR: {total_matched:,} ({coverage_pct}%)")
    print(f"\nMatch methods:")
    for k, v in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    print(f"\nEDGAR parent breakdown:")
    for edgar, count in sorted(edgar_breakdown.items(), key=lambda x: -x[1]):
        print(f"  {edgar}: {count}")
    print(f"\nSaved to: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
